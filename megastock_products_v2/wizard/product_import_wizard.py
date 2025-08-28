# -*- coding: utf-8 -*-

import base64
import io
import logging
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

_logger = logging.getLogger(__name__)


class ProductImportWizard(models.TransientModel):
    _name = 'product.import.wizard'
    _description = 'Wizard para Importación de Productos desde Excel'

    name = fields.Char(string='Nombre', default='Importar Productos desde Excel')
    
    excel_file = fields.Binary(
        string='Archivo Excel',
        required=True,
        help='Archivo Excel con los datos de productos a importar'
    )
    
    filename = fields.Char(string='Nombre del Archivo')
    
    import_mode = fields.Selection([
        ('create_only', 'Solo Crear'),
        ('update_only', 'Solo Actualizar'),
        ('create_update', 'Crear o Actualizar'),
    ], string='Modo de Importación', default='create_update', required=True)
    
    update_existing = fields.Boolean(
        string='Actualizar Existentes',
        default=True,
        help='Si está marcado, actualizará productos existentes basándose en el código interno'
    )
    
    validate_data = fields.Boolean(
        string='Validar Datos',
        default=True,
        help='Validar formato de datos antes de importar'
    )
    
    results = fields.Text(string='Resultados', readonly=True)
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('done', 'Completado'),
        ('error', 'Error')
    ], string='Estado', default='draft')

    @api.model
    def get_template_fields(self):
        """Retorna los campos de la plantilla Excel"""
        return {
            # Campos básicos obligatorios
            'name': 'Nombre del Producto',
            'default_code': 'Código Interno/SKU',
            'categ_id': 'Categoría (nombre)',
            'megastock_category': 'Categoría MEGASTOCK',
            'list_price': 'Precio de Venta',
            'standard_price': 'Costo',
            
            # Dimensiones
            'largo_cm': 'Largo (cm)',
            'ancho_cm': 'Ancho (cm)', 
            'alto_cm': 'Alto (cm)',
            'ceja_cm': 'Ceja (cm)',
            
            # Especificaciones técnicas
            'flauta': 'Flauta (C/B/E)',
            'test_value': 'Test (200/250/275/300)',
            'kl_value': 'KL (32/44)',
            'material_type': 'Material (kraft/interstock/monus/westrock)',
            'colors_printing': 'Colores Impresión (0/1/2/3/4)',
            'gramaje': 'Gramaje (90/125/150/175/200)',
            'tipo_caja': 'Tipo de Caja (tapa_fondo/jumbo/exportacion/americana)',
            
            # Campos adicionales MEGASTOCK
            'numero_troquel': 'Número de Troquel',
            'empaque': 'Empaque (caja/pallet/bulto/unidad/rollo)',
            
            # Campos adicionales
            'uom_id': 'Unidad de Medida',
            'uom_po_id': 'Unidad de Compra',
            'description': 'Descripción',
            'barcode': 'Código de Barras',
            'weight': 'Peso (kg)',
            'volume': 'Volumen (m³)',
            
            # Campos de control
            'active': 'Activo (Verdadero/Falso)',
            'sale_ok': 'Se puede vender',
            'purchase_ok': 'Se puede comprar',
        }

    def action_download_template(self):
        """Genera y descarga la plantilla Excel"""
        if not HAS_XLSXWRITER:
            # Alternativa: crear archivo CSV
            return self._download_csv_template()
            
        try:
            # Crear un libro de trabajo Excel
            from io import BytesIO
            
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            
            # Crear hoja principal
            worksheet = workbook.add_worksheet('Productos')
            
            # Formatos
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#366092',
                'font_color': 'white',
                'border': 1
            })
            
            # Escribir encabezados
            fields = self.get_template_fields()
            row = 0
            for col, (field_name, field_label) in enumerate(fields.items()):
                worksheet.write(row, col, field_label, header_format)
                worksheet.write_comment(row, col, f'Campo técnico: {field_name}')
                
            # Ajustar ancho de columnas
            for col in range(len(fields)):
                worksheet.set_column(col, col, 20)
            
            # Crear hoja de ejemplos
            example_sheet = workbook.add_worksheet('Ejemplos')
            
            # Encabezados de ejemplo
            for col, (field_name, field_label) in enumerate(fields.items()):
                example_sheet.write(0, col, field_label, header_format)
            
            # Datos de ejemplo
            examples = [
                {
                    'name': 'Caja Corrugada 40x30x20',
                    'default_code': 'CAJ-40x30x20-C',
                    'categ_id': 'CAJAS',
                    'megastock_category': 'cajas',
                    'list_price': 2.50,
                    'standard_price': 1.80,
                    'largo_cm': 40.0,
                    'ancho_cm': 30.0,
                    'alto_cm': 20.0,
                    'ceja_cm': 2.0,
                    'flauta': 'c',
                    'test_value': '275',
                    'kl_value': '32',
                    'material_type': 'kraft',
                    'colors_printing': '0',
                    'gramaje': '125',
                    'tipo_caja': 'americana',
                    'numero_troquel': 'TR-001',
                    'empaque': 'caja',
                    'uom_id': 'Unidades',
                    'uom_po_id': 'Unidades',
                    'description': 'Caja de cartón corrugado',
                    'barcode': '',
                    'weight': 0.5,
                    'volume': 0.024,
                    'active': 'Verdadero',
                    'sale_ok': 'Verdadero',
                    'purchase_ok': 'Verdadero'
                },
                {
                    'name': 'Lámina Corrugada 100x70',
                    'default_code': 'LAM-100x70-B',
                    'categ_id': 'LÁMINAS',
                    'megastock_category': 'laminas',
                    'list_price': 5.00,
                    'standard_price': 3.50,
                    'largo_cm': 100.0,
                    'ancho_cm': 70.0,
                    'flauta': 'b',
                    'test_value': '200',
                    'material_type': 'interstock',
                    'colors_printing': '2',
                    'active': 'Verdadero',
                    'sale_ok': 'Verdadero',
                    'purchase_ok': 'Verdadero'
                }
            ]
            
            for row_idx, example in enumerate(examples, 1):
                for col, field_name in enumerate(fields.keys()):
                    value = example.get(field_name, '')
                    example_sheet.write(row_idx, col, value)
            
            # Crear hoja de validaciones
            validation_sheet = workbook.add_worksheet('Validaciones')
            validations = [
                ['Campo', 'Valores Permitidos', 'Descripción'],
                ['megastock_category', 'cajas, laminas, papel, planchas, separadores, materias_primas', 'Categoría MEGASTOCK'],
                ['flauta', 'c, b, e', 'Tipo de flauta del cartón'],
                ['test_value', '200, 250, 275, 300', 'Valor de test del cartón'],
                ['kl_value', '32, 44', 'Kilolibras del material'],
                ['material_type', 'kraft, interstock, monus, westrock', 'Tipo de material'],
                ['colors_printing', '0, 1, 2, 3, 4', 'Número de colores de impresión'],
                ['gramaje', '90, 125, 150, 175, 200', 'Gramaje del papel en g/m²'],
                ['tipo_caja', 'tapa_fondo, jumbo, exportacion, americana', 'Tipo de caja'],
                ['active', 'Verdadero, Falso', 'Producto activo o inactivo'],
                ['sale_ok', 'Verdadero, Falso', 'Permite ventas'],
                ['purchase_ok', 'Verdadero, Falso', 'Permite compras']
            ]
            
            for row_idx, validation in enumerate(validations):
                for col_idx, cell_value in enumerate(validation):
                    if row_idx == 0:
                        validation_sheet.write(row_idx, col_idx, cell_value, header_format)
                    else:
                        validation_sheet.write(row_idx, col_idx, cell_value)
            # Cerrar workbook y obtener datos
            workbook.close()
            excel_data = output.getvalue()
            output.close()
            
            # Crear attachment
            attachment = self.env['ir.attachment'].create({
                'name': 'plantilla_productos_megastock.xlsx',
                'type': 'binary',
                'datas': base64.b64encode(excel_data),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'res_model': self._name,
                'res_id': self.id,
            })
            
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'new',
            }
            
        except Exception as e:
            raise UserError(_('Error al generar plantilla: %s') % str(e))

    def _download_csv_template(self):
        """Método alternativo para generar plantilla CSV si no hay xlsxwriter"""
        try:
            import csv
            from io import StringIO
            
            output = StringIO()
            writer = csv.writer(output)
            
            # Escribir encabezados
            fields = self.get_template_fields()
            headers = list(fields.values())
            writer.writerow(headers)
            
            # Escribir ejemplos
            examples = [
                ['Caja Corrugada 40x30x20', 'CAJ-40x30x20-C', 'CAJAS', 'cajas', 2.50, 1.80, 40.0, 30.0, 20.0, 2.0, 'c', '275', '32', 'kraft', '0', '125', 'americana', 'Unidades', 'Unidades', 'Caja de cartón corrugado', '', 0.5, 0.024, 'Verdadero', 'Verdadero', 'Verdadero'],
                ['Lámina Corrugada 100x70', 'LAM-100x70-B', 'LÁMINAS', 'laminas', 5.00, 3.50, 100.0, 70.0, '', '', 'b', '200', '', 'interstock', '2', '150', '', 'Hojas', 'Hojas', 'Lámina de cartón corrugado', '', 0.8, '', 'Verdadero', 'Verdadero', 'Verdadero']
            ]
            
            for example in examples:
                writer.writerow(example)
            
            # Crear attachment
            csv_data = output.getvalue()
            attachment = self.env['ir.attachment'].create({
                'name': 'plantilla_productos_megastock.csv',
                'type': 'binary',
                'datas': base64.b64encode(csv_data.encode('utf-8')),
                'mimetype': 'text/csv',
                'res_model': self._name,
                'res_id': self.id,
            })
            
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'new',
            }
            
        except Exception as e:
            raise UserError(_('Error al generar plantilla CSV: %s') % str(e))

    def action_import_products(self):
        """Importa productos desde el archivo Excel"""
        if not self.excel_file:
            raise UserError(_('Debe seleccionar un archivo Excel'))
        
        try:
            # Decodificar archivo
            file_data = base64.b64decode(self.excel_file)
            file_obj = io.BytesIO(file_data)
            
            # Leer archivo Excel
            data = []
            
            # Intentar con openpyxl primero (.xlsx)
            if HAS_OPENPYXL:
                try:
                    workbook = openpyxl.load_workbook(file_obj)
                    sheet = workbook.active
                    
                    # Obtener encabezados
                    headers = []
                    for col in range(1, sheet.max_column + 1):
                        headers.append(sheet.cell(row=1, column=col).value)
                    
                    # Leer datos
                    for row in range(2, sheet.max_row + 1):
                        row_data = {}
                        for col in range(1, len(headers) + 1):
                            cell_value = sheet.cell(row=row, column=col).value
                            if headers[col-1]:
                                row_data[headers[col-1]] = cell_value
                        if any(row_data.values()):  # Si hay datos en la fila
                            data.append(row_data)
                    
                except Exception as openpyxl_error:
                    if HAS_XLRD:
                        # Intentar con xlrd para archivos .xls
                        try:
                            file_obj.seek(0)
                            workbook = xlrd.open_workbook(file_contents=file_data)
                            sheet = workbook.sheet_by_index(0)
                            
                            # Obtener encabezados
                            headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                            
                            # Leer datos
                            for row in range(1, sheet.nrows):
                                row_data = {}
                                for col in range(sheet.ncols):
                                    if headers[col]:
                                        row_data[headers[col]] = sheet.cell_value(row, col)
                                if any(str(v).strip() for v in row_data.values() if v is not None):
                                    data.append(row_data)
                        except Exception as xlrd_error:
                            raise UserError(_('No se pudo leer el archivo Excel. Errores:\nopenpyxl: %s\nxlrd: %s') % (openpyxl_error, xlrd_error))
                    else:
                        raise UserError(_('No se pudo leer el archivo Excel: %s\nInstale la librería xlrd para archivos .xls') % openpyxl_error)
            
            elif HAS_XLRD:
                # Solo xlrd disponible
                try:
                    file_obj.seek(0)
                    workbook = xlrd.open_workbook(file_contents=file_data)
                    sheet = workbook.sheet_by_index(0)
                    
                    # Obtener encabezados
                    headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                    
                    # Leer datos
                    for row in range(1, sheet.nrows):
                        row_data = {}
                        for col in range(sheet.ncols):
                            if headers[col]:
                                row_data[headers[col]] = sheet.cell_value(row, col)
                        if any(str(v).strip() for v in row_data.values() if v is not None):
                            data.append(row_data)
                except Exception as e:
                    raise UserError(_('No se pudo leer el archivo Excel con xlrd: %s') % str(e))
            
            else:
                raise UserError(_('No hay librerías disponibles para leer archivos Excel.\nInstale openpyxl y/o xlrd'))
            
            # Si no hay datos, intentar leer como CSV
            if not data:
                try:
                    file_obj.seek(0)
                    content = file_data.decode('utf-8')
                    
                    # Detectar separador
                    import csv
                    dialect = csv.Sniffer().sniff(content[:1024])
                    separator = dialect.delimiter
                    
                    # Leer CSV
                    lines = content.strip().split('\n')
                    if lines:
                        headers = lines[0].split(separator)
                        for i, line in enumerate(lines[1:], 1):
                            if line.strip():
                                values = line.split(separator)
                                row_data = {}
                                for j, header in enumerate(headers):
                                    if j < len(values) and header:
                                        row_data[header] = values[j]
                                if any(str(v).strip() for v in row_data.values() if v):
                                    data.append(row_data)
                                    
                except Exception as csv_error:
                    if not data:  # Solo mostrar error si no hay datos de Excel
                        raise UserError(_('No se pudo leer el archivo como Excel o CSV: %s') % str(csv_error))
            
            # Procesar datos
            results = self._process_excel_data(data)
            
            self.results = results
            self.state = 'done'
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': self._name,
                'res_id': self.id,
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'new',
                'context': self.env.context
            }
            
        except Exception as e:
            self.state = 'error'
            self.results = f'Error durante la importación: {str(e)}'
            _logger.error(f'Error importando productos: {e}')
            raise UserError(_('Error al procesar archivo: %s') % str(e))

    def _process_excel_data(self, data):
        """Procesa los datos del Excel y crea/actualiza productos"""
        results = []
        created_count = 0
        updated_count = 0
        error_count = 0
        
        # Mapeo de campos Excel a campos del modelo
        field_mapping = self._get_field_mapping()
        
        for idx, row_data in enumerate(data, 1):
            try:
                # Preparar valores del producto
                product_vals = self._prepare_product_vals(row_data, field_mapping)
                
                if not product_vals.get('name'):
                    results.append(f'Fila {idx}: Error - Nombre del producto requerido')
                    error_count += 1
                    continue
                
                # Buscar producto existente
                existing_product = None
                if product_vals.get('default_code'):
                    existing_product = self.env['product.template'].search([
                        ('default_code', '=', product_vals['default_code'])
                    ], limit=1)
                
                if existing_product:
                    if self.import_mode in ['update_only', 'create_update']:
                        # Actualizar producto existente
                        existing_product.write(product_vals)
                        updated_count += 1
                        results.append(f'Fila {idx}: Producto actualizado - {product_vals["name"]}')
                    else:
                        results.append(f'Fila {idx}: Producto ya existe - {product_vals["name"]} (omitido)')
                else:
                    if self.import_mode in ['create_only', 'create_update']:
                        # Crear nuevo producto
                        new_product = self.env['product.template'].create(product_vals)
                        created_count += 1
                        results.append(f'Fila {idx}: Producto creado - {new_product.name} ({new_product.default_code})')
                    else:
                        results.append(f'Fila {idx}: Producto no encontrado - {product_vals["name"]} (omitido)')
                
            except Exception as e:
                error_count += 1
                results.append(f'Fila {idx}: Error - {str(e)}')
                _logger.error(f'Error procesando fila {idx}: {e}')
        
        # Resumen
        summary = f'''
RESUMEN DE IMPORTACIÓN:
========================
Productos creados: {created_count}
Productos actualizados: {updated_count}  
Errores: {error_count}
Total procesado: {len(data)} filas

DETALLES:
=========
'''
        
        return summary + '\n'.join(results)

    def _get_field_mapping(self):
        """Retorna el mapeo entre campos de Excel y campos del modelo"""
        return {
            'Nombre del Producto': 'name',
            'Código Interno/SKU': 'default_code',
            'Categoría (nombre)': 'categ_id',
            'Categoría MEGASTOCK': 'megastock_category',
            'Precio de Venta': 'list_price',
            'Costo': 'standard_price',
            'Largo (cm)': 'largo_cm',
            'Ancho (cm)': 'ancho_cm',
            'Alto (cm)': 'alto_cm',
            'Ceja (cm)': 'ceja_cm',
            'Flauta (C/B/E)': 'flauta',
            'Test (200/250/275/300)': 'test_value',
            'KL (32/44)': 'kl_value',
            'Material (kraft/interstock/monus/westrock)': 'material_type',
            'Colores Impresión (0/1/2/3/4)': 'colors_printing',
            'Gramaje (90/125/150/175/200)': 'gramaje',
            'Tipo de Caja (tapa_fondo/jumbo/exportacion/americana)': 'tipo_caja',
            'Unidad de Medida': 'uom_id',
            'Unidad de Compra': 'uom_po_id',
            'Descripción': 'description',
            'Código de Barras': 'barcode',
            'Peso (kg)': 'weight',
            'Volumen (m³)': 'volume',
            'Activo (Verdadero/Falso)': 'active',
            'Se puede vender': 'sale_ok',
            'Se puede comprar': 'purchase_ok',
        }

    def _prepare_product_vals(self, row_data, field_mapping):
        """Prepara los valores del producto desde los datos de Excel"""
        vals = {}
        
        for excel_field, odoo_field in field_mapping.items():
            if excel_field in row_data and row_data[excel_field] is not None:
                value = row_data[excel_field]
                
                # Procesar valores según el campo
                if odoo_field == 'categ_id':
                    # Buscar categoría por nombre
                    if value:
                        category = self.env['product.category'].search([('name', '=', value)], limit=1)
                        if category:
                            vals[odoo_field] = category.id
                        else:
                            # Crear categoría si no existe
                            category = self.env['product.category'].create({'name': str(value)})
                            vals[odoo_field] = category.id
                
                elif odoo_field in ['uom_id', 'uom_po_id']:
                    # Buscar unidad de medida por nombre
                    if value:
                        uom = self.env['uom.uom'].search([('name', '=', value)], limit=1)
                        if uom:
                            vals[odoo_field] = uom.id
                
                elif odoo_field in ['active', 'sale_ok', 'purchase_ok']:
                    # Convertir valores booleanos
                    if isinstance(value, str):
                        vals[odoo_field] = value.lower() in ['verdadero', 'true', 'sí', 'si', '1', 'yes']
                    else:
                        vals[odoo_field] = bool(value)
                
                elif odoo_field in ['list_price', 'standard_price', 'largo_cm', 'ancho_cm', 'alto_cm', 'ceja_cm', 'weight', 'volume']:
                    # Convertir valores numéricos
                    try:
                        vals[odoo_field] = float(value) if value else 0.0
                    except (ValueError, TypeError):
                        vals[odoo_field] = 0.0
                
                elif odoo_field in ['test_value', 'kl_value', 'gramaje', 'colors_printing']:
                    # Campos de selección numéricos
                    if value:
                        try:
                            # Convertir a entero para eliminar decimales, luego a string
                            numeric_value = int(float(str(value)))
                            vals[odoo_field] = str(numeric_value)
                        except (ValueError, TypeError):
                            vals[odoo_field] = str(value) if value else False
                    else:
                        vals[odoo_field] = False
                
                elif odoo_field in ['flauta', 'material_type', 'tipo_caja', 'megastock_category']:
                    # Campos de selección de texto
                    vals[odoo_field] = str(value).lower() if value else False
                
                elif odoo_field == 'barcode':
                    # Manejo especial para códigos de barras (pueden venir en notación científica)
                    if value:
                        try:
                            # Si es número en notación científica, convertir a entero
                            if 'E' in str(value).upper() or 'e' in str(value):
                                numeric_value = int(float(str(value)))
                                vals[odoo_field] = str(numeric_value)
                            else:
                                vals[odoo_field] = str(value)
                        except (ValueError, TypeError):
                            vals[odoo_field] = str(value) if value else ''
                    else:
                        vals[odoo_field] = ''
                        
                else:
                    # Campos de texto simples
                    vals[odoo_field] = str(value) if value else ''
        
        # Valores por defecto
        if 'active' not in vals:
            vals['active'] = True
        if 'sale_ok' not in vals:
            vals['sale_ok'] = True
        if 'purchase_ok' not in vals:
            vals['purchase_ok'] = True
            
        return vals